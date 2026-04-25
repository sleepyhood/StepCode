import os
import sys
import json
import time
import requests
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from urllib.parse import quote, urlparse
from datetime import datetime

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.keys import Keys
    from webdriver_manager.chrome import ChromeDriverManager
    HAS_SELENIUM = True
except ImportError:
    HAS_SELENIUM = False


try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class ScoreExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("StepCode 성적 추출기 (대학생 공지용)")
        self.root.geometry("800x850")
        
        self.stop_event = threading.Event()
        self.is_running = False
        
        self._build_ui()
        
        if not HAS_OPENPYXL:
            messagebox.showwarning("라이브러리 누락", "openpyxl 라이브러리가 설치되어 있지 않아 엑셀 생성이 불가능할 수 있습니다.\n`pip install openpyxl`을 실행해주세요.")
    
    def _build_ui(self):
        # 1. 서버 및 인증 정보
        auth_frame = tk.LabelFrame(self.root, text="[ 1. 서버 및 계정 정보 ]", padx=10, pady=10)
        auth_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(auth_frame, text="서버 주소:").grid(row=0, column=0, sticky="e")
        self.base_url_var = tk.StringVar(value="http://edu.doingcoding.com")
        tk.Entry(auth_frame, textvariable=self.base_url_var, width=40).grid(row=0, column=1, sticky="w", padx=5)
        
        tk.Label(auth_frame, text="관리자 ID:").grid(row=1, column=0, sticky="e", pady=5)
        self.admin_id_var = tk.StringVar()
        tk.Entry(auth_frame, textvariable=self.admin_id_var, width=20).grid(row=1, column=1, sticky="w", padx=5)
        
        tk.Label(auth_frame, text="관리자 PW:").grid(row=1, column=2, sticky="e", pady=5)
        self.admin_pw_var = tk.StringVar()
        tk.Entry(auth_frame, textvariable=self.admin_pw_var, show="*", width=20).grid(row=1, column=3, sticky="w", padx=5)
        
        # 2. 문제 구간 설정
        prob_frame = tk.LabelFrame(self.root, text="[ 2. 문제 범위 설정 ]", padx=10, pady=10)
        prob_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(prob_frame, text="문제 접두사(Prefix):").grid(row=0, column=0, sticky="e")
        self.prefix_var = tk.StringVar(value="ky_2026_01_26682_")
        tk.Entry(prob_frame, textvariable=self.prefix_var, width=25).grid(row=0, column=1, sticky="w", padx=5)
        
        tk.Label(prob_frame, text="시작 번호:").grid(row=0, column=2, sticky="e")
        self.start_id_var = tk.IntVar(value=1)
        tk.Entry(prob_frame, textvariable=self.start_id_var, width=5).grid(row=0, column=3, sticky="w", padx=5)
        
        tk.Label(prob_frame, text="~ 끝 번호:").grid(row=0, column=4, sticky="e")
        self.end_id_var = tk.IntVar(value=10)
        tk.Entry(prob_frame, textvariable=self.end_id_var, width=5).grid(row=0, column=5, sticky="w", padx=5)
        
        tk.Label(prob_frame, text="자리수(Padding):").grid(row=1, column=0, sticky="e", pady=5)
        self.pad_var = tk.IntVar(value=2)
        tk.Entry(prob_frame, textvariable=self.pad_var, width=5).grid(row=1, column=1, sticky="w", padx=5)
        tk.Label(prob_frame, text="(예: 2로 설정 시 01, 02...)", fg="gray").grid(row=1, column=2, columnspan=2, sticky="w")
        
        # 3. 대상 학생 명단
        student_frame = tk.LabelFrame(self.root, text="[ 3. 대상 학생 명단 (줄바꿈으로 구분) ]", padx=10, pady=10)
        student_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.student_text = scrolledtext.ScrolledText(student_frame, height=8)
        self.student_text.pack(fill="both", expand=True)
        self.student_text.insert(tk.END, "student_01\nstudent_02\nstudent_03") # 기본 더미 데이터
        
        # 4. 저장 및 실행
        exec_frame = tk.LabelFrame(self.root, text="[ 4. 저장 경로 및 실행 ]", padx=10, pady=10)
        exec_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(exec_frame, text="저장 폴더:").grid(row=0, column=0, sticky="e")
        self.save_dir_var = tk.StringVar(value=os.getcwd())
        tk.Entry(exec_frame, textvariable=self.save_dir_var, width=50, state="readonly").grid(row=0, column=1, sticky="w", padx=5)
        tk.Button(exec_frame, text="변경", command=self._select_dir).grid(row=0, column=2)
        
        self.mock_mode_var = tk.BooleanVar(value=True) # 기본값을 True로 설정 (테스트용)
        self.mock_cb = tk.Checkbutton(exec_frame, text="테스트 모드 (네트워크 접속 안 함)", 
                                    variable=self.mock_mode_var, command=self._on_mock_mode_toggle)
        self.mock_cb.grid(row=1, column=0, sticky="w", pady=5)
        
        self.use_browser_var = tk.BooleanVar(value=False) # 테스트 모드와 모순되지 않게 기본값 False
        self.use_browser_cb = tk.Checkbutton(exec_frame, text="브라우저 자동 로그인 사용 (Selenium)", 
                                           variable=self.use_browser_var)
        self.use_browser_cb.grid(row=1, column=1, sticky="w", pady=5)
        
        self.headless_var = tk.BooleanVar(value=False)
        self.headless_cb = tk.Checkbutton(exec_frame, text="브라우저 창 숨기기 (Headless)", 
                                        variable=self.headless_var)
        self.headless_cb.grid(row=1, column=2, sticky="w", pady=5)
        
        # 엑셀 문제 제목 포함 옵션 추가 (기본값 False)
        self.include_titles_var = tk.BooleanVar(value=False)
        self.include_titles_cb = tk.Checkbutton(exec_frame, text="엑셀에 문제 제목 포함 (API 추가 호출)", 
                                              variable=self.include_titles_var)
        self.include_titles_cb.grid(row=2, column=0, columnspan=3, sticky="w", pady=5)
        
        # 초기 상태 설정 (테스트 모드면 브라우저 옵션 비활성화)
        self._on_mock_mode_toggle()

        btn_frame = tk.Frame(exec_frame)
        btn_frame.grid(row=3, column=0, columnspan=3, pady=10)
        
        self.start_btn = tk.Button(btn_frame, text="🚀 추출 시작", font=("Helvetica", 12, "bold"), bg="#4CAF50", fg="black", command=self._start_extraction)
        self.start_btn.pack(side="left", padx=10)
        
        self.stop_btn = tk.Button(btn_frame, text="🛑 중단", font=("Helvetica", 12, "bold"), bg="#f44336", fg="black", command=self._stop_extraction, state=tk.DISABLED)
        self.stop_btn.pack(side="left")
        
        # 로그 창
        self.log_area = scrolledtext.ScrolledText(self.root, height=10, state="disabled", bg="#f0f0f0")
        self.log_area.pack(fill="both", expand=True, padx=10, pady=10)
        
    def _select_dir(self):
        d = filedialog.askdirectory(initialdir=self.save_dir_var.get())
        if d:
            self.save_dir_var.set(d)
            
    def _on_mock_mode_toggle(self):
        """테스트 모드 활성화 여부에 따라 브라우저 관련 옵션 제어"""
        if self.mock_mode_var.get():
            # 테스트 모드일 때: 브라우저 옵션 비활성화 및 해제
            self.use_browser_cb.config(state=tk.DISABLED)
            self.headless_cb.config(state=tk.DISABLED)
            self.use_browser_var.set(False)
            self.headless_var.set(False)
        else:
            # 일반 모드일 때: 브라우저 옵션 활성화
            self.use_browser_cb.config(state=tk.NORMAL)
            self.headless_cb.config(state=tk.NORMAL)
            
    def log(self, msg):
        self.log_area.config(state="normal")
        self.log_area.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_area.see(tk.END)
        self.log_area.config(state="disabled")
        self.root.update()

    def _start_extraction(self):
        if self.is_running:
            return
            
        students_raw = self.student_text.get("1.0", tk.END).strip()
        if not students_raw:
            messagebox.showwarning("오류", "학생 명단을 입력해주세요.")
            return
            
        self.student_list = [s.strip() for s in students_raw.split("\n") if s.strip()]
        
        self.is_running = True
        self.stop_event.clear()
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.student_text.config(state=tk.DISABLED)
        
        self.log("="*50)
        self.log(f"추출 시작... 총 대상: {len(self.student_list)}명")
        
        threading.Thread(target=self._extraction_task, daemon=True).start()

    def _stop_extraction(self):
        self.stop_event.set()
        self.log("중단 요청됨. 진행 중인 작업을 마무리하는 중...")

    def _extraction_task(self):
        base_url = self.base_url_var.get().strip()
        admin_id = self.admin_id_var.get().strip()
        admin_pw = self.admin_pw_var.get().strip()
        prefix = self.prefix_var.get().strip()
        start_id = self.start_id_var.get()
        end_id = self.end_id_var.get()
        pad = self.pad_var.get()
        is_mock = self.mock_mode_var.get()
        save_dir = self.save_dir_var.get()
        
        use_browser = self.use_browser_var.get()
        headless = self.headless_var.get()
        
        # 문제 ID 리스트 생성
        problem_ids = [f"{prefix}{str(i).zfill(pad)}" for i in range(start_id, end_id + 1)]
        self.log(f"대상 문제: {problem_ids[0]} ~ {problem_ids[-1]} ({len(problem_ids)}개)")
        
        if not is_mock:
            self.log("인증 시도 중...")
            if use_browser and HAS_SELENIUM:
                session = self._login_selenium(base_url, admin_id, admin_pw, headless)
            else:
                if use_browser and not HAS_SELENIUM:
                    self.log("⚠️ Selenium이 설치되어 있지 않아 일반 요청 방식으로 시도합니다.")
                session = self._login(base_url, admin_id, admin_pw)
                
            if not session:
                self.log("❌ 로그인 실패. 중단합니다.")
                self._finish_task()
                return
            self.log("✅ 로그인 성공! (세션 획득)")
        else:
            self.log("⚠️ 테스트 모드로 실행 (가짜 데이터 생성)")
            session = None
            
        # 문제 제목 가져오기 연동
        self.title_map = {}
        if self.include_titles_var.get() and session:
            self.title_map = self._fetch_problem_titles(base_url, session, problem_ids)
            
        all_data = []
        error_students = []
        
        # Checkpoint load
        checkpoint_file = os.path.join(save_dir, "scores_checkpoint.json")
        processed_users = set()
        if os.path.exists(checkpoint_file):
            try:
                with open(checkpoint_file, "r", encoding="utf-8") as f:
                    all_data = json.load(f)
                processed_users = {d["username"] for d in all_data}
                self.log(f"🔄 체크포인트 로드 됨: 이미 완료된 {len(processed_users)}명 건너뜀.")
            except:
                self.log("⚠️ 체크포인트 파일을 읽을 수 없어 새로 시작합니다.")
                
        print(f"[분석] 추출 대상 문제 목록: {problem_ids}")                
        # Main Loop
        for i, username in enumerate(self.student_list):
            if self.stop_event.is_set():
                break
                
            if username in processed_users:
                continue
                
            self.log(f"조회 중 ({i+1}/{len(self.student_list)}): {username}")
            
            try:
                if is_mock:
                    time.sleep(0.05) # 모의 지연
                    import random
                    student_scores = {"username": username}
                    # 20% 확률로 미제출, 나머지는 무작위 점수
                    for pid in problem_ids:
                        if random.random() > 0.2:
                            student_scores[pid] = random.choice([0, 50, 80, 100, 100, 100])
                    all_data.append(student_scores)
                else:
                    res = session.get(f"{base_url}/api/profile?username={quote(username)}", timeout=10)
                    data = res.json()
                    print(data)
                    if data.get("error"):
                        self.log(f"  ! Error: {data['error']}")
                        error_students.append(username)
                    else:
                        problems = data.get("data", {}).get("oi_problems_status", {}).get("problems", {})
                        student_scores = {"username": username}
                        for internal_id, p_info in problems.items():
                            public_id = p_info.get("_id") # 실제 문제 번호인 'P101v0101'을 꺼냄
                            # 👇 이 줄을 추가해서 GUI에서 생성된 ID 목록과 실제 ID를 눈으로 비교해 보세요.
                            print(f"[진단] 실제 ID: '{public_id}' | 내가 찾는 ID 목록에 있는가?: {public_id in problem_ids}")
                            print(f"[진단] 실제 ID: '{public_id}' (타입: {type(public_id)})")
                            print(f"[진단] 대상 목록의 첫 번째 값: '{problem_ids[0]}' (타입: {type(problem_ids[0])})")

                            if public_id in problem_ids: # 'P101v0101'과 'P101v0101'을 비교해서 성공!
                                student_scores[public_id] = p_info.get("score", 0)
                        all_data.append(student_scores)
                    time.sleep(0.15)
                    
            except Exception as e:
                self.log(f"  ! Exception: {e}")
                error_students.append(username)
                
            # Checkpoint save
            if (i + 1) % 50 == 0:
                self._save_checkpoint(checkpoint_file, all_data)
                
        # Final Processing
        self.log("✅ 조회 루프 완료. 엑셀 생성을 준비합니다.")
        self._save_checkpoint(checkpoint_file, all_data) # final
        
        if error_students:
            err_file = os.path.join(save_dir, "error_students.log")
            with open(err_file, "w", encoding="utf-8") as f:
                f.write("\n".join(error_students))
            self.log(f"⚠️ 실패한 학생 {len(error_students)}명은 error_students.log에 저장되었습니다.")
            
        self._generate_excel(all_data, problem_ids, save_dir, getattr(self, 'title_map', {}))
        self._finish_task()
        
    def _login_selenium(self, base_url, username, password, headless):
        """Selenium을 이용한 브라우저 기반 로그인 및 세션 획득 (login.py 방식)"""
        if not HAS_SELENIUM:
            self.log("❌ Selenium 라이브러리가 없습니다.")
            return None
            
        self.log(f"브라우저 실행 중 (Headless={headless})...")
        driver = None
        try:
            options = Options()
            if headless:
                options.add_argument("--headless")
            
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("useAutomationExtension", False)
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

            # ChromeDriver 자동 관리 및 실행
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            
            # 봇 감지 우회를 위한 스크립트 실행
            driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            })

            # 1. 메인 페이지 접속
            self.log(f"메인 페이지 접속 중: {base_url}")
            driver.get(base_url)
            time.sleep(3) # Vue 컴포넌트 로딩 대기
            
            # 2. 로그인 버튼 클릭 (사용자 제공 최신 XPath 적용)
            self.log("로그인 버튼 탐색 중...")
            try:
                # 1순위: 사용자 제공 정확한 경로
                # 2순위: 유연한 텍스트/클래스 탐색
                login_btn = None
                try:
                    login_btn = driver.find_element(By.XPATH, '//*[@id="header"]/div[1]/div[2]/div[1]/button[1]')
                except:
                    login_btn = driver.find_element(By.XPATH, "//button[contains(text(), '로그인') or contains(text(), 'Login')]")
                
                login_btn.click()
                time.sleep(1.5) # 팝업 애니메이션 대기
            except Exception as e:
                self.log(f"⚠️ 로그인 버튼 자동 클릭 실패: {e}")
                if not headless:
                    self.log("창에서 직접 로그인 버튼을 클릭해 주세요. (30초 대기)")
                    time.sleep(30)
                else:
                    raise Exception("로그인 버튼 탐색 실패")

            # 3. 로그인 정보 입력
            self.log("로그인 정보 입력 중...")
            try:
                # 팝업 내 입력창 (id_input, pw_input은 구조가 고정적이므로 상대경로 권장하나 일단 기존 로직 유지)
                id_input = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/div/div[2]/div/form/div[1]/div/div[1]/input')
                id_input.send_keys(username)
                
                pw_input = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/div/div[2]/div/form/div[2]/div/div/input')
                pw_input.send_keys(password)
                
                # 4. 로그인 제출
                self.log("인증 요청 중...")
                pw_input.send_keys(Keys.ENTER)
                time.sleep(3) # 인증 처리 및 페이지 전환 대기

                # 로그인이 완료되어 쿠키가 구워질 때까지 최대 10초 대기
                self.log("인증 완료 대기 중...")
                for _ in range(10):
                    cookies = driver.get_cookies()
                    if any(c['name'] == 'csrftoken' for c in cookies):
                        break
                    time.sleep(1)
            except Exception as e:
                self.log("⚠️ 자동 폼 입력 실패. 브라우저 창에서 수동 로그인이 필요합니다.")
                if not headless:
                    time.sleep(60)

            # 5. 세션 추출 및 requests 연결 (CSRF 헤더 포함)
            self.log("인증 세션 추출 및 CSRF 설정 중...")
            cookies = driver.get_cookies()
            session = requests.Session()
            session.headers.update({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Referer": base_url
            })
            
            csrf_token = None
            for cookie in cookies:
                session.cookies.set(cookie['name'], cookie['value'], domain=cookie.get('domain'))
                if cookie['name'] == 'csrftoken':
                    csrf_token = cookie['value']
            
            if csrf_token:
                session.headers.update({"X-CSRFToken": csrf_token})
                self.log("✅ CSRF 토큰 설정 완료.")
            else:
                self.log("⚠️ CSRF 토큰을 찾지 못했습니다. 요청이 거부될 수 있습니다.")
            
            driver.quit()
            return session

        except Exception as e:
            self.log(f"❌ 브라우저 로그인 에러: {e}")
            if driver:
                driver.quit()
            return None

    def _login(self, base_url, username, password):
        """CSRF 토큰을 처리하는 정밀 로그인 로직"""
        session = requests.Session()
        # 브라우저처럼 보이기 위한 기본 헤더
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        
        try:
            self.log("초기 보안 토큰(CSRF) 획득 중...")
            # 1. 초기 토큰 획득을 위한 GET 요청
            session.get(base_url, timeout=10) 
            csrf_token = session.cookies.get('csrftoken')
            
            if not csrf_token:
                self.log("⚠️ 메인 페이지에서 CSRF 토큰을 찾을 수 없습니다. 로그인 페이지에서 재시도합니다.")
                session.get(f"{base_url}/login", timeout=10)
                csrf_token = session.cookies.get('csrftoken')

            if not csrf_token:
                self.log("❌ 보안 토큰(CSRF)을 가져오지 못했습니다. 서버 상태나 도메인을 확인하세요.")
                return None

            self.log("✅ 보안 토큰 획득 완료. 인증을 시도합니다.")

            # 2. 헤더 구성
            headers = {
                "X-CSRFToken": csrf_token,
                "Referer": f"{base_url}/login",
                "Content-Type": "application/json"
            }
            
            # 3. 로그인 시도
            login_data = {"username": username, "password": password}
            res = session.post(f"{base_url}/api/login", json=login_data, headers=headers, timeout=10)
            
            # 4. 결과 확인
            if res.status_code == 200:
                try:
                    res_json = res.json()
                    if res_json.get("error") is None:
                        return session
                    else:
                        self.log(f"❌ 로그인 거부 (API 에러): {res_json.get('data', '알 수 없는 오류')}")
                except Exception:
                    self.log("❌ 로그인 응답이 정상적인 JSON 형식이 아닙니다.")
            else:
                self.log(f"❌ 로그인 실패: HTTP {res.status_code}")
                
        except requests.RequestException as e:
            self.log(f"❌ 네트워크 통신 오류: {e}")
        except Exception as e:
            self.log(f"❌ 로그인 프로세스 예외 발생: {e}")
        
        return None

    def _fetch_problem_titles(self, base_url, session, problem_ids):
        """
        API를 통해 지정된 범위 내 문제들의 제목을 가져옵니다.
        """
        self.log("[분석] 문제 제목 정보를 가져오는 중...")
        title_map = {}
        target_set = set(problem_ids)
        page = 1
        limit = 100
        
        try:
            while target_set:
                params = {
                    "paging": "true",
                    "offset": (page - 1) * limit,
                    "limit": limit,
                    "page": page,
                }
                res = session.get(f"{base_url}/api/problem", params=params, timeout=10)
                res.raise_for_status()
                
                data = res.json() or {}
                batch = (data.get("data") or {}).get("results") or []
                
                if not batch:
                    break # 더 이상 페이지가 없으면 종료
                    
                for item in batch:
                    pid = item.get("_id")
                    title = item.get("title")
                    if pid in target_set:
                        title_map[pid] = title
                        target_set.remove(pid) # 찾은 문제는 타겟에서 제거
                
                page += 1
                time.sleep(0.1) # 서버 부하 방지
                
            self.log(f"✅ {len(title_map)}개의 문제 제목을 확보했습니다.")
        except Exception as e:
            self.log(f"⚠️ 문제 제목 가져오기 실패: {e}")
            
        return title_map

    def _save_checkpoint(self, path, data):
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.log("💾 체크포인트 저장 완료.")
        except Exception as e:
            self.log(f"⚠️ 체크포인트 저장 실패: {e}")

    def _generate_excel(self, data, problem_ids, save_dir, title_map=None):
        if not HAS_OPENPYXL:
            self.log("❌ openpyxl이 없어 엑셀을 생성할 수 없습니다.")
            return
            
        if title_map is None:
            title_map = {}
            
        self.log("📊 데이터 집계 및 엑셀 포맷팅 시작...")
        
        # Calculate stats
        for row in data:
            total = 0
            count = 0
            feedback = []
            for pid in problem_ids:
                val = row.get(pid, "미제출")
                if isinstance(val, (int, float)):
                    total += val
                    count += 1
                    if val < 50:
                        feedback.append(f"{pid[-2:]}번 미흡")
                else:
                    feedback.append(f"{pid[-2:]}번 미제출")
                    
            row["Total"] = total
            row["Average"] = total / len(problem_ids) if problem_ids else 0
            row["Feedback"] = ", ".join(feedback) if feedback else "통과"
            
        # Rank sorting
        data.sort(key=lambda x: x["Total"], reverse=True)
        for i, row in enumerate(data):
            row["Rank"] = i + 1
            
        # Generate XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Scores"
        
        # 헤더 생성: title_map에 제목이 있으면 병기
        headers = ["Rank", "username"]
        for pid in problem_ids:
            title = title_map.get(pid)
            if title:
                headers.append(f"{pid}\n({title})")
            else:
                headers.append(pid)
        headers.extend(["Total", "Average", "Feedback"])
        
        ws.append(headers)
        
        # Style headers
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # 텍스트 길이에 따라 적절히 너비 조절
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
        ws.row_dimensions[1].height = 30 # 줄바꿈된 헤더를 위해 높이 지정
        
        # Data & Missing styles
        missing_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for row_idx, row_data in enumerate(data, 2):
            # 1. Rank & Username (1, 2열)
            ws.cell(row=row_idx, column=1, value=row_data.get("Rank", ""))
            ws.cell(row=row_idx, column=2, value=row_data.get("username", ""))
            
            # 2. Problems (3열부터 시작, 원래 ID로 데이터 조회)
            for j, pid in enumerate(problem_ids):
                val = row_data.get(pid, "미제출")
                cell = ws.cell(row=row_idx, column=3+j, value=val)
                if val == "미제출":
                    cell.fill = missing_fill
            
            # 3. Stats (문제 열 종료 후 순차 기록)
            base_col = 3 + len(problem_ids)
            ws.cell(row=row_idx, column=base_col, value=row_data.get("Total", 0))
            ws.cell(row=row_idx, column=base_col + 1, value=row_data.get("Average", 0))
            ws.cell(row=row_idx, column=base_col + 2, value=row_data.get("Feedback", ""))
                    
        # Conditional Formatting (Gradient)
        rule = ColorScaleRule(start_type='num', start_value=0, start_color='FF6347', # Red
                              mid_type='num', mid_value=50, mid_color='FFD700',      # Yellow
                              end_type='num', end_value=100, end_color='90EE90')     # Green
                              
        # 점수 열과 합계/평균 열에 색상 적용
        # 문제 열 범위: 3 ~ 2+len(problem_ids)
        # 통계 열 범위: base_col ~ base_col+1 (Total, Average)
        for col_idx in range(3, base_col + 2):
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{len(data)+1}', rule)
                
        output_path = os.path.join(save_dir, "student_scores.xlsx")
        try:
            wb.save(output_path)
            self.log(f"✅ 엑셀 파일 저장 성공: {output_path}")
        except Exception as e:
            self.log(f"❌ 엑셀 저장 에러: {e}")

    def _finish_task(self):
        self.is_running = False
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.student_text.config(state=tk.NORMAL)
        self.log("="*50)
        self.log("작업이 완전히 종료되었습니다.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ScoreExtractorApp(root)
    root.mainloop()
