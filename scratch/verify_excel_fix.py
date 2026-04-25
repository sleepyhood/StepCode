
import os
import sys
import json
import tkinter as tk
from unittest.mock import MagicMock

# StepCode 경로 추가
sys.path.append(r'c:\Users\osw\Desktop\Workspace\Projects\StepCode\Resources\tools')

try:
    from extract_scores import ScoreExtractorApp
    import openpyxl
    from openpyxl import Workbook
except ImportError as e:
    print(f"Import Error: {e}")
    sys.exit(1)

def test_excel_logic():
    print("🚀 Excel 데이터 매핑 로직 검증 시작...")
    
    # 실재 Tkinter 루트 (GUI 변수 생성을 위해 필요)
    root = tk.Tk()
    root.withdraw() # 창은 숨김
    app = ScoreExtractorApp(root)
    
    # 1. 테스트 데이터 준비
    problem_ids = ["P101v0101", "P101v0102"]
    title_map = {"P101v0101": "기초1", "P101v0102": "기초2"}
    
    # row_data에는 '원래 ID'가 키로 들어있음
    data = [
        {
            "Rank": 1,
            "username": "tester",
            "P101v0101": 100,
            "P101v0102": 90,
            "Total": 190,
            "Average": 95,
            "Feedback": "통과"
        }
    ]
    
    save_dir = r"c:\Users\osw\Desktop\Workspace\Projects\StepCode\scratch"
    os.makedirs(save_dir, exist_ok=True)
    
    # 2. 메서드 실행
    print("📦 _generate_excel 실행 중...")
    app._generate_excel(data, problem_ids, save_dir, title_map)
    
    # 3. 파일 생성 확인 및 내용 검증
    xl_path = os.path.join(save_dir, "student_scores.xlsx")
    if not os.path.exists(xl_path):
        print("❌ 실패: 엑셀 파일이 생성되지 않았습니다.")
        return False
        
    wb = openpyxl.load_workbook(xl_path)
    ws = wb.active
    
    # 헤더 검증
    h3 = ws.cell(row=1, column=3).value
    print(f"헤더 3열: {repr(h3)}")
    if "P101v0101" not in h3 or "기초1" not in h3:
        print("❌ 실패: 헤더에 제목이 포함되지 않았습니다.")
        return False

    # 데이터 검증 (가장 중요한 부분)
    v3 = ws.cell(row=2, column=3).value
    v4 = ws.cell(row=2, column=4).value
    print(f"데이터 3열(P101v0101 점수): {v3}")
    print(f"데이터 4열(P101v0102 점수): {v4}")
    
    if v3 == 100 and v4 == 90:
        print("✅ 성공: 데이터가 원래 ID와 매핑되어 정확히 기록되었습니다!")
        return True
    else:
        print(f"❌ 실패: 데이터 매핑 오류 (값: {v3}, {v4})")
        return False

if __name__ == "__main__":
    success = test_excel_logic()
    if not success:
        sys.exit(1)
