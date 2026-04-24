import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def generate_mock_data():
    return [
        {"username": "student_01", "prob_01": 100, "prob_02": 95, "prob_03": 0},
        {"username": "student_02", "prob_01": 80, "prob_02": "미제출", "prob_03": 20},
        {"username": "student_03", "prob_01": 100, "prob_02": 100, "prob_03": 100},
        {"username": "student_04", "prob_01": "미제출", "prob_02": "미제출", "prob_03": "미제출"},
    ]

def main():
    data = generate_mock_data()
    prob_keys = ["prob_01", "prob_02", "prob_03"]
    
    # Calculate Total and Average
    for row in data:
        total = 0
        count = 0
        feedback = []
        for k in prob_keys:
            val = row.get(k, "미제출")
            if isinstance(val, (int, float)):
                total += val
                count += 1
                if val < 50:
                    feedback.append(f"{k} 낮음")
            else:
                feedback.append(f"{k} 미제출")
        
        row["Total"] = total
        row["Average"] = total / len(prob_keys) if len(prob_keys) > 0 else 0
        row["Feedback"] = ", ".join(feedback) if feedback else "통과"

    # Sort by total for Rank
    data.sort(key=lambda x: x["Total"], reverse=True)
    for i, row in enumerate(data):
        row["Rank"] = i + 1

    # Restore sorting to original or just keep sorted by Rank
    # We will keep it sorted by Rank

    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    
    headers = ["Rank", "username"] + prob_keys + ["Total", "Average", "Feedback"]
    ws.append(headers)
    
    # Format headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Append data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if header in prob_keys:
                if val == "미제출":
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Add Color Scale (Red - Yellow - Green) to problem columns and Total/Average
    # Red: FF6347, Yellow: FFD700, Green: 90EE90
    rule = ColorScaleRule(start_type='num', start_value=0, start_color='FF6347',
                          mid_type='num', mid_value=50, mid_color='FFD700',
                          end_type='num', end_value=100, end_color='90EE90')
    
    # Apply to problem columns
    for col_idx, header in enumerate(headers, 1):
        if header in prob_keys:
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{len(data)+1}', rule)
    
    wb.save("scratch/test_scores.xlsx")
    print("Excel generation successful. Columns:", headers)

if __name__ == "__main__":
    main()
